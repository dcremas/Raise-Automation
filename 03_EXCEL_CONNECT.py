import boto3
from openpyxl import load_workbook
from datetime import datetime, date

date_tod = date.today().strftime("%Y-%m-%d")
year_tod, month_tod, day_tod = date_tod[:4], date_tod[5:7], date_tod[8:]

excel_file_source = f'FILES_EXCEL/data_{month_tod}-{day_tod}-{year_tod}.xlsx'
excel_tab_source = 'Data'

excel_file_dest = 'RAISE_SYW_2025-CURRENT.xlsx'
excel_tab_dest = 'Data'

try:
	workbook_source = load_workbook(excel_file_source)
	workbook_dest = load_workbook(excel_file_dest)

	sheet_source = workbook_source[excel_tab_source]
	sheet_dest = workbook_dest[excel_tab_dest]

	for idx, record in enumerate(sheet_source.iter_rows()):
		if idx == 0:
			continue
		sheet_dest[f'A{idx+1}'].value = record[0].value
		sheet_dest[f'B{idx+1}'].value = record[1].value
		sheet_dest[f'C{idx+1}'].value = record[2].value
		sheet_dest[f'D{idx+1}'].value = record[3].value
	
	workbook_dest.save(excel_file_dest)

	s3_client = boto3.client('s3')
	S3_BUCKET_NAME = 'ec2-sandbox-dcremas'
	S3_STAGING_PREFIX = 'daily_files_excel/'
	s3_key = f"{S3_STAGING_PREFIX}RAISE-SYW_{month_tod}-{day_tod}-{year_tod}.xlsx"
	with open(excel_file_dest, 'rb') as f:
		s3_response = s3_client.put_object(
			Body=f,
			Bucket=S3_BUCKET_NAME,
			Key=s3_key,
			ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
			)
    
	print(f"Excel file {excel_file_dest} updated successfully.")

except Exception as e:
	print(f"An error occured: {e}")
